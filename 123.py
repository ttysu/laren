#! python3
#-*- coding: utf-8 -*
import os
import openpyxl

os.chdir('E:\\预算')
#修改路径
mz = []#定义空列表
mz = os.listdir()
nl = []
for i in range(len(mz)-1):
    wb = openpyxl.load_workbook(mz[i],data_only=True)
    sheet = wb.get_sheet_by_name('汇总')
    A1 = sheet.cell(row=3,column=1).value
    A2 = sheet.cell(row=3,column=5).value
    A3 = sheet.cell(row=3,column=6).value
    A4 = sheet.cell(row=3, column=7).value
    nl.append(A1)
    nl.append(A2)
    nl.append(A3)
    nl.append(A4)
    wb.save(mz[i])

print(nl)
end='D'+str(len(mz))
i=0
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
for rowOfCellObjects in sheet['A1':'D4']:
    for cellObj in rowOfCellObjects:
        cellObj.value=nl[i]
        i=i+1
wb.save('abc.xlsx')
