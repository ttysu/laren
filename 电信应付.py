#! python3
# -*- coding:utf-8 -*-
import os
import openpyxl

os.chdir('E:\\基础表\电信')
# 修改路径
mz = []  # 定义空列表
mz = os.listdir()
nl = []
for i in range(len(mz)):
    os.chdir('E:\\基础表\电信')
    wb = openpyxl.load_workbook(mz[i], data_only=True)
    sheet = wb.get_sheet_by_name('应付')
    A1 = sheet.cell(row=2, column=1).value#工程名
    A2 = sheet.cell(row=3, column=1).value#站点名
    A3 = sheet.cell(row=3, column=5).value#折扣
    A4 = sheet.cell(row=3, column=7).value#面积
    A5 = sheet.cell(row=55, column=6).value  # 工程概、预算总费用
    A6 = sheet.cell(row=56, column=6).value # 让利（建安费-材料费）*（1-折扣）
    A7 = sheet.cell(row=57, column=6).value  # 集成费（不含材料费）
    wb.save(mz[i])
    os.chdir('E:\\')
    wb = openpyxl.load_workbook('abc.xlsx')
    sheet = wb.get_active_sheet()
    sheet.cell(row=i + 1, column=1).value = A1
    sheet.cell(row=i + 1, column=2).value = A2
    sheet.cell(row=i + 1, column=3).value = A3
    sheet.cell(row=i + 1, column=4).value = A4
    sheet.cell(row=i + 1, column=5).value = A5
    sheet.cell(row=i + 1, column=6).value = A6
    sheet.cell(row=i + 1, column=7).value = A7
    wb.save('abc.xlsx')