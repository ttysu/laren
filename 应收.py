#! python3
# -*- coding:utf-8 -*-
import openpyxl
import os
os.chdir('E:\\基础表\预算设计')
# 修改路径
mz = []  # 定义空列表
mz = os.listdir()
nl = []
for i in range(len(mz)):
    os.chdir('E:\\基础表\预算设计')
    wb = openpyxl.load_workbook(mz[i], data_only=True)
    sheet = wb.get_sheet_by_name('汇总')
    A1 =A2=A3=A4=A5=A6=A7=A8=A9=A10=A11=A12=A13= None
    A1 = sheet.cell(row=2, column=1).value  # 工程名
    A2 = sheet.cell(row=3, column=1).value  # 站点名
    A3 = sheet.cell(row=3, column=5).value  # 折扣
    A4 = sheet.cell(row=3, column=7).value  # 面积
    print(A2)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == '让利（人工+措施+企业管理+利润+仪表+机械）*（1-折扣）':
                A5 = sheet.cell(row=cell.row, column=6).value # 让利（人工+措施+企业管理+利润+仪表+机械）*（1-折扣）
                print(cell.row, cell.value)
            elif cell.value == '集成费除税价（不含税金，不含材料费）' :
                A6 = sheet.cell(row=cell.row, column=6).value # 集成费除税价（不含税金，不含材料费）
                print(cell.row, cell.value)
            elif cell.value == '集成费增值税（税率11%）':
                A7 = sheet.cell(row=cell.row, column=6).value #集成费增值税（税率11%）
                print(cell.row, cell.value)
            elif cell.value == '材料费除税价':
                A8 = sheet.cell(row=cell.row, column=6).value #材料费除税价
                print(cell.row, cell.value)
            elif cell.value == '材料费增值税（税率11%）' :
                A9 = sheet.cell(row=cell.row, column=6).value  # 材料费增值税（税率11%）
                print(cell.row, cell.value)
            elif cell.value == '安全生产费除税价':
                A10 = sheet.cell(row=cell.row, column=6).value  # 安全生产费除税价
                print(cell.row, cell.value)
            elif cell.value == '安全生产费增值税' or cell.value == '安全生产费增值税（税率11%）':
                A11 = sheet.cell(row=cell.row, column=6).value  # 安全生产费增值税
                print(cell.row, cell.value)
            elif cell.value == '合计' or cell.value == '收入总计' or cell.value == '收入合计':
                A12 = sheet.cell(row=cell.row, column=6).value  # 合计
                print(cell.row, cell.value)
            elif cell.value == '集成费（不含材料费）':
                A13 = sheet.cell(row=cell.row, column=6).value  # 合计
                print(cell.row, cell.value)

    wb.save(mz[i])
    os.chdir('E:\\')
    wb = openpyxl.load_workbook('abc.xlsx')
    sheet = wb.get_active_sheet()
    sheet.cell(row=i + 1, column=1).value = A1
    sheet.cell(row=i + 1, column=2).value = A2
    sheet.cell(row=i + 1, column=3).value = A3
    sheet.cell(row=i + 1, column=4).value = A4
    sheet.cell(row=i + 1, column=5).value = A5
    sheet.cell(row=i + 1, column=6).value = A6
    sheet.cell(row=i + 1, column=7).value = A7
    sheet.cell(row=i + 1, column=8).value = A8
    sheet.cell(row=i + 1, column=9).value = A9
    sheet.cell(row=i + 1, column=10).value = A10
    sheet.cell(row=i + 1, column=11).value = A11
    sheet.cell(row=i + 1, column=12).value = A12
    sheet.cell(row=i + 1, column=13).value = A13
    wb.save('abc.xlsx')
