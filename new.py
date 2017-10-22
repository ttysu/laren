#! python3
# -*- coding:utf-8 -*-
import os
import openpyxl

os.chdir('E:\\预算1')
# 修改路径
mz = []  # 定义空列表
mz = os.listdir()
nl = []
for i in range(len(mz)):
    os.chdir('E:\\预算1')
    wb = openpyxl.load_workbook(mz[i], data_only=True)
    sheet = wb.get_sheet_by_name('汇总')
    A1 = sheet.cell(row=3, column=1).value
    A2 = sheet.cell(row=3, column=5).value
    A3 = sheet.cell(row=3, column=6).value
    A4 = sheet.cell(row=3, column=7).value
    wb.save(mz[i])
    os.chdir('E:\\')
    wb = openpyxl.load_workbook('abc.xlsx')
    sheet = wb.get_active_sheet()
    sheet.cell(row=i + 1, column=1).value = A1
    sheet.cell(row=i + 1, column=2).value = A2
    sheet.cell(row=i + 1, column=3).value = A3
    sheet.cell(row=i + 1, column=4).value = A4
    wb.save('abc.xlsx')

